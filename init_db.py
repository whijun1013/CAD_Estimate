import openpyxl
import os
import re
import argparse
import logging
from datetime import datetime
from dotenv import load_dotenv
from database import engine, SessionLocal, Base
from models import Project, ApartmentType, MaterialSpecification, HardwareSpecification, CabinetBOM, BuildingQuantity, CabinetPriceMaster

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# Excel Column & Sheet Constants
SHEET_PROJECT_INFO = "현장정보"
SHEET_SPEC_PREFIX = "사양서"
SHEET_BOM_PREFIX = "내역서"
SHEET_DONG_PREFIX = "동정보"

# Rows and Columns configuration
INFO_SCAN_MAX_ROW = 40
INFO_SCAN_MAX_COL = 10
APT_TYPE_START_ROW = 10
APT_TYPE_END_ROW = 30
APT_TYPE_COL_CODE = 11
APT_TYPE_COL_NAME = 12
APT_TYPE_COL_QTY = 13
APT_TYPE_COL_CHANGE = 14

SPEC_START_ROW = 4
SPEC_MAX_ROW = 100
SPEC_COL_PART = 4
SPEC_COL_THICKNESS = 6
SPEC_COL_GRADE = 7
SPEC_COL_MATERIAL = 8
SPEC_COL_FINISH = 9
SPEC_COL_GRAIN = 10
SPEC_COL_PRIMARY_MAT = 11
SPEC_COL_PRIMARY_MAT_DTL = 12
SPEC_COL_BACKING = 13
SPEC_COL_BACKING_DTL = 14
SPEC_COL_EDGE = 15
SPEC_COL_EDGE_DTL = 16

HW_START_ROW = 8
HW_MAX_ROW = 100
HW_COL_GRP = 19
HW_COL_NAME = 20
HW_COL_APP = 21
HW_COL_REM = 22

BOM_START_ROW = 4
BOM_MAX_ROW = 250

DONG_START_ROW = 8
DONG_MAX_ROW = 250
DONG_START_COL = 14

def parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def clean_int(val, context=""):
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    s = str(val).strip().replace(",", "")
    if not s:
        return 0
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            logging.warning("Failed to parse integer value '%s' at context: %s. Defaulting to 0.", val, context)
            return 0

def seed_price_masters(db):
    logging.info("Seeding CabinetPriceMaster standard prices...")

    # Standard prices for common items
    default_prices = [
        # (category, product_name, unit_price, product_code)
        ("상부장", "상부장", 75000, "W-TOP-STD"),
        ("상부장", "냉장고장 상부 플랩장", 78000, "W-FLAP-REF"),
        ("하부장", "하부장", 95000, "B-BASE-STD"),
        ("키큰장", "키큰장", 120000, "T-TALL-STD"),
        ("피라/앤드판넬", "마감 판넬", 45000, "P-END-STD"),
        ("피라/앤드판넬", "좌측 마감 판넬 (일반)", 45000, "P-END-LEFT"),
        ("피라/앤드판넬", "우측 마감 판넬 (비규격)", 58000, "P-END-RIGHT-SPC"),
        ("피라/앤드판넬", "앤드판넬", 45000, "P-END-GEN"),
        ("코니스/걸레받이", "상부 마감 휠라 (코니스)", 28000, "C-FILLA-TOP"),
        ("코니스/걸레받이", "코니스", 28000, "C-CORNICE"),
        ("코니스/걸레받이", "휠라", 25000, "C-FILLA"),
        ("코니스/걸레받이", "걸레받이", 22000, "C-PLINTH")
    ]

    for cat, name, price, code in default_prices:
        existing = db.query(CabinetPriceMaster).filter(CabinetPriceMaster.product_name == name).first()
        if not existing:
            pm = CabinetPriceMaster(
                category=cat,
                product_name=name,
                unit_price=price,
                product_code=code,
                remarks="Default seed standard price"
            )
            db.add(pm)
    db.commit()
    logging.info("CabinetPriceMaster seeding complete.")

def init_database(reset=False, excel_path=None):
    if reset:
        logging.info("Dropping all existing database tables for fresh seed...")
        Base.metadata.drop_all(bind=engine)
        logging.info("Creating database tables...")
        Base.metadata.create_all(bind=engine)
    else:
        logging.info("Ensuring database tables exist (no drop)...")
        Base.metadata.create_all(bind=engine)

    db = SessionLocal()
    seed_price_masters(db)

    if not excel_path:
        excel_path = os.getenv("EXCEL_PATH", "sample/PO_BR_262603000301_0_1.xlsx")

    if not os.path.exists(excel_path):
        logging.error("Excel file not found at %s! Skipping Excel import.", excel_path)
        db.close()
        return

    logging.info("Opening Excel file: %s", excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    try:
        # 1. Parse Project Info from '현장정보'
        sheet_info = wb[SHEET_PROJECT_INFO]

        info_dict = {}
        for r in range(1, INFO_SCAN_MAX_ROW + 1):
            for c in range(1, INFO_SCAN_MAX_COL + 1):
                cell_val = sheet_info.cell(row=r, column=c).value
                if cell_val:
                    cell_val = str(cell_val).strip()
                    if cell_val in ("거래선", "현장명", "품목", "현장  주소", "현장담당자", "현장연락처",
                                    "시공거래선", "시공담당자", "시공연락처", "최초투입일", "입주/오픈예정일",
                                    "현장종류", "최고층높이", "분절공사", "PO번호", "계약번호", "작성완료일"):
                        val = sheet_info.cell(row=r, column=c+1).value
                        if val is None or val == "":
                            val = sheet_info.cell(row=r, column=c+2).value
                            if val is None or val == "":
                                val = sheet_info.cell(row=r, column=c+3).value

                        if val is not None:
                            info_dict[cell_val] = str(val).strip()

        logging.info("Extracted Project Info:")
        for k, v in info_dict.items():
            logging.info("  %s: %s", k, v)

        po_number = info_dict.get("PO번호", "26-2603-0003-01")

        # Check if project already exists
        existing_project = db.query(Project).filter(Project.po_number == po_number).first()
        if existing_project:
            logging.info("Project with PO Number %s already exists. Skipping project creation.", po_number)
            project = existing_project
        else:
            project = Project(
                po_number=po_number,
                contract_number=info_dict.get("계약번호"),
                name=info_dict.get("현장명", "김해삼계푸르지오"),
                client=info_dict.get("거래선"),
                partner_installer=info_dict.get("시공거래선"),
                item_type=info_dict.get("품목"),
                address=info_dict.get("현장  주소"),
                manager_name=info_dict.get("현장담당자"),
                manager_contact=info_dict.get("현장연락처"),
                installer_name=info_dict.get("시공담당자"),
                installer_contact=info_dict.get("시공연락처"),
                first_delivery_date=parse_date(info_dict.get("최초투입일")),
                opening_date=parse_date(info_dict.get("입주/오픈예정일")),
                site_type=info_dict.get("현장종류"),
                max_floor=clean_int(info_dict.get("최고층높이"), "최고층높이"),
                is_divided_work=(info_dict.get("분절공사") == "Y"),
                remarks="Imported from PO excel file."
            )
            db.add(project)
            db.commit()
            db.refresh(project)
            logging.info("Created Project: %s (ID: %d)", project.name, project.id)

        # Parse Apartment Types summary
        apartment_types_dict = {}
        for r in range(APT_TYPE_START_ROW, APT_TYPE_END_ROW + 1):
            t_code = sheet_info.cell(row=r, column=APT_TYPE_COL_CODE).value
            t_name = sheet_info.cell(row=r, column=APT_TYPE_COL_NAME).value
            t_qty = sheet_info.cell(row=r, column=APT_TYPE_COL_QTY).value
            t_change = sheet_info.cell(row=r, column=APT_TYPE_COL_CHANGE).value

            if t_code and t_name and str(t_code).strip() != "합계":
                t_code = str(t_code).strip()
                t_name = str(t_name).strip()
                qty = clean_int(t_qty, f"Apartment Type Qty for {t_code}")
                is_changed = (str(t_change).strip() == "Y")

                apartment_types_dict[t_code] = {
                    "name": t_name,
                    "qty": qty,
                    "is_changed": is_changed
                }

        type_models = {}
        for t_code, info in apartment_types_dict.items():
            existing_type = db.query(ApartmentType).filter(
                ApartmentType.project_id == project.id,
                ApartmentType.type_name == t_code
            ).first()

            if existing_type:
                type_models[t_code] = existing_type
            else:
                apt_type = ApartmentType(
                    project_id=project.id,
                    type_name=t_code,
                    household_count=info["qty"],
                    is_changed=info["is_changed"]
                )
                db.add(apt_type)
                db.commit()
                db.refresh(apt_type)
                type_models[t_code] = apt_type

        # 2. Parse Specs, BOM, and Dong Info
        for t_code, apt_type in type_models.items():
            logging.info("Processing details for type: %s...", t_code)

            # --- Parse Specifications ---
            spec_sheet_name = f"{SHEET_SPEC_PREFIX}({t_code})"
            if spec_sheet_name in wb.sheetnames:
                spec_sheet = wb[spec_sheet_name]

                # Check if specs already seeded to prevent duplication without reset
                spec_count = db.query(MaterialSpecification).filter(MaterialSpecification.type_id == apt_type.id).count()
                if spec_count == 0:
                    current_category = ""
                    for r in range(SPEC_START_ROW, SPEC_MAX_ROW + 1):
                        part_val = spec_sheet.cell(row=r, column=SPEC_COL_PART).value
                        if part_val:
                            part_val = str(part_val).strip()
                            if part_val.startswith("[[") and part_val.endswith("]]"):
                                current_category = part_val.replace("[[", "").replace("]]", "")
                                continue

                            if current_category and part_val in ("몸통", "문짝", "코니스", "앤드판넬", "휠라", "걸레받이"):
                                thickness = spec_sheet.cell(row=r, column=SPEC_COL_THICKNESS).value
                                grade = spec_sheet.cell(row=r, column=SPEC_COL_GRADE).value
                                material = spec_sheet.cell(row=r, column=SPEC_COL_MATERIAL).value
                                finish_method = spec_sheet.cell(row=r, column=SPEC_COL_FINISH).value
                                grain_direction = spec_sheet.cell(row=r, column=SPEC_COL_GRAIN).value

                                primary_material = spec_sheet.cell(row=r, column=SPEC_COL_PRIMARY_MAT).value
                                primary_detail = spec_sheet.cell(row=r, column=SPEC_COL_PRIMARY_MAT_DTL).value

                                backing_material = spec_sheet.cell(row=r, column=SPEC_COL_BACKING).value
                                backing_detail = spec_sheet.cell(row=r, column=SPEC_COL_BACKING_DTL).value

                                edge_material = spec_sheet.cell(row=r, column=SPEC_COL_EDGE).value
                                edge_detail = spec_sheet.cell(row=r, column=SPEC_COL_EDGE_DTL).value

                                spec = MaterialSpecification(
                                    type_id=apt_type.id,
                                    category=current_category,
                                    part_name=part_val,
                                    thickness=str(thickness).strip() if thickness else None,
                                    grade=str(grade).strip() if grade else None,
                                    material=str(material).strip() if material else None,
                                    finish_method=str(finish_method).strip() if finish_method else None,
                                    grain_direction=str(grain_direction).strip() if grain_direction else None,
                                    primary_material=str(primary_material).strip() if primary_material else None,
                                    primary_material_detail=str(primary_detail).strip() if primary_detail else None,
                                    backing_material=str(backing_material).strip() if backing_material else None,
                                    backing_material_detail=str(backing_detail).strip() if backing_detail else None,
                                    edge_material=str(edge_material).strip() if edge_material else None,
                                    edge_material_detail=str(edge_detail).strip() if edge_detail else None
                                )
                                db.add(spec)

                    # Hardware Specs
                    for r in range(HW_START_ROW, HW_MAX_ROW + 1):
                        item_grp = spec_sheet.cell(row=r, column=HW_COL_GRP).value
                        item_name = spec_sheet.cell(row=r, column=HW_COL_NAME).value
                        app_val = spec_sheet.cell(row=r, column=HW_COL_APP).value
                        rem_val = spec_sheet.cell(row=r, column=HW_COL_REM).value

                        if item_name and str(item_name).strip():
                            item_name = str(item_name).strip()
                            item_grp = str(item_grp).strip() if item_grp else "사양"

                            hw_spec = HardwareSpecification(
                                type_id=apt_type.id,
                                item_group=item_grp,
                                item_name=item_name,
                                application=str(app_val).strip() if app_val else None,
                                special_remarks=str(rem_val).strip() if rem_val else None
                            )
                            db.add(hw_spec)

                    db.commit()
                    logging.info("  Specs and hardware seeded.")
                else:
                    logging.info("  Specs already exist for %s. Skipping spec seed.", t_code)

            # --- Parse BOM ---
            bom_sheet_name = f"{SHEET_BOM_PREFIX}({t_code})"
            bom_items_by_no_name = {}

            if bom_sheet_name in wb.sheetnames:
                bom_sheet = wb[bom_sheet_name]

                bom_exist_count = db.query(CabinetBOM).filter(CabinetBOM.type_id == apt_type.id).count()
                if bom_exist_count == 0:
                    current_category = ""
                    bom_count = 0
                    for r in range(BOM_START_ROW, BOM_MAX_ROW + 1):
                        col1 = bom_sheet.cell(row=r, column=1).value
                        col5_no = bom_sheet.cell(row=r, column=5).value
                        col6_name = bom_sheet.cell(row=r, column=6).value

                        if col1 and str(col1).strip() in ("상부장", "하부장", "키큰장", "보조주방", "보조주방1", "보조주방2"):
                            current_category = str(col1).strip()
                            continue

                        no_val = clean_int(col5_no, f"BOM Item No row {r}")
                        if no_val > 0 and col6_name:
                            status = bom_sheet.cell(row=r, column=1).value
                            is_special = bom_sheet.cell(row=r, column=4).value
                            product_code = bom_sheet.cell(row=r, column=7).value
                            attr_code = bom_sheet.cell(row=r, column=8).value

                            w = bom_sheet.cell(row=r, column=9).value
                            h = bom_sheet.cell(row=r, column=10).value
                            d = bom_sheet.cell(row=r, column=11).value

                            direction = bom_sheet.cell(row=r, column=12).value
                            qd_l = bom_sheet.cell(row=r, column=13).value
                            qd_m = bom_sheet.cell(row=r, column=14).value
                            qd_r = bom_sheet.cell(row=r, column=15).value
                            qo_l = bom_sheet.cell(row=r, column=16).value
                            qo_m = bom_sheet.cell(row=r, column=17).value
                            qo_r = bom_sheet.cell(row=r, column=18).value
                            qty_sum = bom_sheet.cell(row=r, column=19).value
                            remarks = bom_sheet.cell(row=r, column=20).value

                            bom_item = CabinetBOM(
                                type_id=apt_type.id,
                                category=current_category if current_category else "기타",
                                status=str(status).strip() if status else None,
                                is_special=(str(is_special).strip() == "S"),
                                item_no=no_val,
                                product_name=str(col6_name).strip(),
                                product_code=str(product_code).strip() if product_code else None,
                                attribute_code=str(attr_code).strip() if attr_code else None,
                                width=clean_int(w, f"BOM width row {r}"),
                                height=clean_int(h, f"BOM height row {r}"),
                                depth=clean_int(d, f"BOM depth row {r}"),
                                base_direction=str(direction).strip() if direction else None,
                                qty_drawing_left=clean_int(qd_l, f"BOM qd_l row {r}"),
                                qty_drawing_mid=clean_int(qd_m, f"BOM qd_m row {r}"),
                                qty_drawing_right=clean_int(qd_r, f"BOM qd_r row {r}"),
                                qty_opposite_left=clean_int(qo_l, f"BOM qo_l row {r}"),
                                qty_opposite_mid=clean_int(qo_m, f"BOM qo_m row {r}"),
                                qty_opposite_right=clean_int(qo_r, f"BOM qo_r row {r}"),
                                qty_sum=clean_int(qty_sum, f"BOM qty_sum row {r}"),
                                remarks=str(remarks).strip() if remarks else None
                            )
                            db.add(bom_item)
                            bom_count += 1
                            db.flush()
                            bom_items_by_no_name[(no_val, str(col6_name).strip())] = bom_item

                    db.commit()
                    logging.info("  Seeded %d BOM line items.", bom_count)
                else:
                    logging.info("  BOM items already exist for %s. Re-linking for Dong Info...", t_code)
                    # Pull existing BOMs to reconstruct the dict
                    existing_boms = db.query(CabinetBOM).filter(CabinetBOM.type_id == apt_type.id).all()
                    for b in existing_boms:
                        bom_items_by_no_name[(b.item_no, b.product_name)] = b

            # --- Parse Dong Info ---
            dong_sheet_name = f"{SHEET_DONG_PREFIX}({t_code})"
            if dong_sheet_name in wb.sheetnames:
                dong_sheet = wb[dong_sheet_name]

                # Check if building quantities seeded
                first_bom = list(bom_items_by_no_name.values())
                bq_exists = False
                if first_bom:
                    bq_exists = db.query(BuildingQuantity).filter(BuildingQuantity.bom_id == first_bom[0].id).count() > 0

                if not bq_exists:
                    col_mappings = {}
                    current_building = ""
                    max_cols = dong_sheet.max_column

                    for c in range(DONG_START_COL, max_cols + 1):
                        header6 = dong_sheet.cell(row=6, column=c).value
                        if not header6:
                            break

                        b_val = dong_sheet.cell(row=4, column=c).value
                        if b_val:
                            current_building = str(b_val).strip()

                        line_val = None
                        for check_c in range(c, DONG_START_COL - 1, -1):
                            lv = dong_sheet.cell(row=5, column=check_c).value
                            if lv is not None:
                                line_val = str(lv).strip()
                                break

                        if str(header6).strip() == "합계" and current_building and line_val:
                            col_mappings[c] = (current_building, line_val)

                    dong_qty_count = 0
                    for r in range(DONG_START_ROW, DONG_MAX_ROW + 1):
                        col4_no = dong_sheet.cell(row=r, column=4).value
                        col6_name = dong_sheet.cell(row=r, column=6).value

                        no_val = clean_int(col4_no, f"Dong Info Row {r}")
                        if no_val > 0 and col6_name:
                            product_name = str(col6_name).strip()
                            bom_item = bom_items_by_no_name.get((no_val, product_name))

                            if bom_item:
                                for c, (b_no, l_no) in col_mappings.items():
                                    qty_val = dong_sheet.cell(row=r, column=c).value
                                    qty = clean_int(qty_val, f"Dong Qty row {r} col {c}")
                                    if qty > 0:
                                        bq = BuildingQuantity(
                                            bom_id=bom_item.id,
                                            building_no=b_no,
                                            line_no=l_no,
                                            qty=qty
                                        )
                                        db.add(bq)
                                        dong_qty_count += 1

                    db.commit()
                    logging.info("  Seeded %d building quantities.", dong_qty_count)
                else:
                    logging.info("  Building quantities already exist for %s. Skipping.", t_code)

        logging.info("All database seeding operations completed successfully.")

    except Exception as ex:
        db.rollback()
        logging.error("Database error occurred: %s", ex)
        raise ex
    finally:
        db.close()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Seed construction order database.")
    parser.add_argument("--reset", action="store_true", help="Reset all tables (drop and recreate)")
    parser.add_argument("--excel", type=str, default=None, help="Path to sample excel file")
    args = parser.parse_args()

    init_database(reset=args.reset, excel_path=args.excel)
