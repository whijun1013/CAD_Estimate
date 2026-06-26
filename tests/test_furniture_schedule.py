import os
import sys
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

# Set up system path to load project modules
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from main import app, get_db
from database import Base
import models

def test_furniture_schedule_endpoint():
    # 1. Setup in-memory SQLite database
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool
    )
    TestingSessionLocal = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)

    # 2. Dependency override for db session
    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    # 3. Seed data
    db = TestingSessionLocal()
    try:
        # Create Project
        proj = models.Project(po_number="PO-SCH-1", name="Schedule Test Site")
        db.add(proj)
        db.commit()

        # Create Apartment Type
        apt_type = models.ApartmentType(project_id=proj.id, type_name="84A", household_count=100)
        db.add(apt_type)
        db.commit()
        db.refresh(apt_type)

        # Create CabinetBOM entries with various scenarios:
        # Scenario A: Full explicit dimensions (no inference needed)
        bom_explicit = models.CabinetBOM(
            type_id=apt_type.id,
            category="상부장",
            item_no=1,
            product_name="900*620*350 후드장",
            product_code="C0001",
            attribute_code="W900",
            width=900,
            height=620,
            depth=350,
            qty_sum=10,
            remarks="도면 텍스트 기반 확정"
        )
        # Scenario B: Missing height and depth, should trigger category inference for "하부장"
        bom_inferred_bottom = models.CabinetBOM(
            type_id=apt_type.id,
            category="하부장",
            item_no=2,
            product_name="600*850 하부장",
            product_code="C0002",
            attribute_code="W600",
            width=600,
            height=0, # missing
            depth=None, # missing
            qty_sum=5,
            remarks="높이/깊이 추론 필요"
        )
        # Scenario C: Missing height and depth, cornice matching
        bom_inferred_cornice = models.CabinetBOM(
            type_id=apt_type.id,
            category="기타",
            item_no=3,
            product_name="걸레받이",
            product_code="C0003",
            attribute_code="BASEBOARD",
            width=1200,
            height=None,
            depth=0,
            qty_sum=2,
            remarks="걸레받이 매칭"
        )

        db.add_all([bom_explicit, bom_inferred_bottom, bom_inferred_cornice])
        db.commit()

        # --- Test cases ---

        # Test Case 1: Fetch entire schedule
        response = client.get(f"/api/apartment-types/{apt_type.id}/furniture-schedule")
        assert response.status_code == 200
        data = response.json()

        assert data["apartment_type"] == "84A"
        items = data["items"]
        assert len(items) == 3

        # Verify Scenario A (Explicit dimensions)
        item1 = next(it for it in items if it["item_no"] == 1)
        assert item1["width_mm"] == 900
        assert item1["height_mm"] == 620
        assert item1["depth_mm"] == 350
        assert item1["spec_label"] == "900*620*350"
        assert item1["qty"] == 10
        assert item1["dimension_source"]["width"] == "drawing_text"
        assert item1["dimension_source"]["height"] == "drawing_text"
        assert item1["dimension_source"]["depth"] == "drawing_text"
        assert item1["confidence"] == 1.0
        assert item1["needs_review"] is False
        assert item1["review_reason"] is None

        # Verify Scenario B (Inferred bottom cabinet: H=850, D=600)
        item2 = next(it for it in items if it["item_no"] == 2)
        assert item2["width_mm"] == 600
        assert item2["height_mm"] == 850
        assert item2["depth_mm"] == 600
        assert item2["spec_label"] == "600*850*600"
        assert item2["qty"] == 5
        assert item2["dimension_source"]["width"] == "drawing_text"
        assert item2["dimension_source"]["height"] == "default_by_category"
        assert item2["dimension_source"]["depth"] == "default_by_category"
        assert item2["confidence"] == 0.75 # two inferred fields
        assert item2["needs_review"] is True
        assert "높이값" in item2["review_reason"]
        assert "깊이값" in item2["review_reason"]

        # Verify Scenario C (Inferred cornice/baseboard: H=80, D=18)
        item3 = next(it for it in items if it["item_no"] == 3)
        assert item3["width_mm"] == 1200
        assert item3["height_mm"] == 80
        assert item3["depth_mm"] == 18
        assert item3["spec_label"] == "1200*80*18"
        assert item3["qty"] == 2
        assert item3["dimension_source"]["width"] == "drawing_text"
        assert item3["dimension_source"]["height"] == "default_by_category"
        assert item3["dimension_source"]["depth"] == "default_by_category"
        assert item3["confidence"] == 0.75
        assert item3["needs_review"] is True
        assert "걸레받이" in item3["review_reason"]

        # Verify summary stats
        assert data["summary"]["total_item_types"] == 3
        assert data["summary"]["total_quantity"] == 17  # 10 + 5 + 2
        assert data["summary"]["review_required_count"] == 2 # item 2 and item 3

        # Test Case 2: Filter by category
        response_cat = client.get(f"/api/apartment-types/{apt_type.id}/furniture-schedule?category=상부장")
        assert response_cat.status_code == 200
        data_cat = response_cat.json()
        assert len(data_cat["items"]) == 1
        assert data_cat["items"][0]["item_no"] == 1

        # Test Case 3: Filter by needs_review = True
        response_rev = client.get(f"/api/apartment-types/{apt_type.id}/furniture-schedule?needs_review=true")
        assert response_rev.status_code == 200
        data_rev = response_rev.json()
        assert len(data_rev["items"]) == 2
        assert {item["item_no"] for item in data_rev["items"]} == {2, 3}

        # Test Case 4: Filter by needs_review = False
        response_norev = client.get(f"/api/apartment-types/{apt_type.id}/furniture-schedule?needs_review=false")
        assert response_norev.status_code == 200
        data_norev = response_norev.json()
        assert len(data_norev["items"]) == 1
        assert data_norev["items"][0]["item_no"] == 1

        # Test Case 5: Filter by search (name)
        response_search = client.get(f"/api/apartment-types/{apt_type.id}/furniture-schedule?search=후드장")
        assert response_search.status_code == 200
        data_search = response_search.json()
        assert len(data_search["items"]) == 1
        assert data_search["items"][0]["item_no"] == 1

        # Test Case 6: Price columns absence check
        # Explicit check: ensure keys like 'price', 'unit_price', 'sum_price' are not in items
        for item in items:
            assert "price" not in item
            assert "unit_price" not in item
            assert "sum_price" not in item

    finally:
        db.close()
        app.dependency_overrides.clear()


def test_furniture_schedule_xlsx_export():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool
    )
    TestingSessionLocal = sessionmaker(bind=engine)
    Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[get_db] = override_get_db
    client = TestClient(app)

    db = TestingSessionLocal()
    try:
        proj = models.Project(po_number="PO-SCH-EXP", name="Export Test Site")
        db.add(proj)
        db.commit()

        apt_type = models.ApartmentType(project_id=proj.id, type_name="84A", household_count=100)
        db.add(apt_type)
        db.commit()
        db.refresh(apt_type)

        bom = models.CabinetBOM(
            type_id=apt_type.id,
            category="상부장",
            item_no=1,
            product_name="후드장",
            product_code="C0001",
            attribute_code="W900",
            width=900,
            height=620,
            depth=350,
            qty_sum=10,
            remarks="도면 텍스트 기반 확정"
        )
        db.add(bom)
        db.commit()

        response = client.get(f"/api/apartment-types/{apt_type.id}/furniture-schedule.xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment; filename" in response.headers["content-disposition"]

        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        assert "도면 분석 요약" in wb.sheetnames
        assert "필요 가구 산출표" in wb.sheetnames

        ws = wb["필요 가구 산출표"]
        assert ws.cell(row=2, column=3).value == "후드장"
        assert ws.cell(row=2, column=5).value == 900

    finally:
        db.close()
        app.dependency_overrides.clear()
