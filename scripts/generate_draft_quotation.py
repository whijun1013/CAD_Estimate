import os
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import datetime
from database import SessionLocal
import models
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def check_dwg_header(dwg_path):
    if not os.path.exists(dwg_path):
        return False, "File not found", 0

    file_size = os.path.getsize(dwg_path)
    try:
        with open(dwg_path, "rb") as f:
            header_bytes = f.read(6)
        if header_bytes.startswith(b"AC10"):
            version_sig = header_bytes.decode("utf-8", errors="ignore")
            return True, f"Valid DWG (Header version: {version_sig})", file_size
        else:
            return False, f"Invalid DWG signature: {header_bytes}", file_size
    except Exception as e:
        return False, f"Read error: {str(e)}", file_size

def generate_quotation(task_id=None, output_path_override=None):
    is_mock = True
    project_name = "Synthetic Public Demo"
    dwg_path = "local_private/private_design.dwg"
    dwg_info = "N/A"
    file_size_mb = 0.0
    limit_status = "N/A"
    raw_items = []

    if task_id is not None:
        import json

        db_sess = SessionLocal()
        try:
            task = db_sess.query(models.CADTask).filter(models.CADTask.id == task_id).first()
            if not task:
                print(f"Error: Task with ID {task_id} not found. Falling back to mock data.")
            else:
                is_mock = False
                project_name = f"Task ID {task.id} - {task.file_name}"
                dwg_path = task.file_path
                # Check header
                is_valid, dwg_info, file_size = check_dwg_header(dwg_path)
                file_size_mb = file_size / (1024 * 1024)
                limit_mb = 50.0
                limit_status = "OK (Below 50MB limit)" if file_size_mb <= limit_mb else "Warning (Above 50MB)"

                if task.structured_analysis:
                    analysis = json.loads(task.structured_analysis)
                    items_list = analysis.get("items", [])
                    for idx, it in enumerate(items_list, 1):
                        raw_items.append({
                            "category": it.get("category") or "기타",
                            "product_name": it.get("product_name") or "가구 품목",
                            "product_code": it.get("product_code") or f"P-CODE-{idx}",
                            "width_mm": it.get("width_mm") or it.get("width") or 0,
                            "depth_mm": it.get("depth_mm") or it.get("depth") or 0,
                            "height_mm": it.get("height_mm") or it.get("height") or 0,
                            "quantity": it.get("quantity") or it.get("qty") or 1,
                            "remarks": it.get("remarks") or ""
                        })
                else:
                    print("Warning: Task has no structured analysis. Falling back to mock data.")
                    is_mock = True
        except Exception as e:
            print(f"Error querying task: {e}. Falling back to mock data.")
            is_mock = True
        finally:
            db_sess.close()

    if is_mock:
        dwg_path = "local_private/private_design.dwg"
        # 1. File verification
        is_valid, dwg_info, file_size = check_dwg_header(dwg_path)
        file_size_mb = file_size / (1024 * 1024)
        limit_mb = 50.0
        limit_status = "OK (Below 50MB limit)" if file_size_mb <= limit_mb else "Warning (Above 50MB)"

        print(f"DWG verification status: {is_valid}")
        print(f"DWG info: {dwg_info}")
        print(f"File size: {file_size_mb:.2f} MB ({limit_status})")

        # 2. Simulated drawing items with missing values to showcase inference engine
        raw_items = [
            {
                "category": "상부장",
                "product_name": "상부장",
                "product_code": "W-TOP-STD",
                "width_mm": 800,
                "depth_mm": 320,
                "height_mm": 700,
                "quantity": 450,
                "remarks": "주방 상부 표준 수납장"
            },
            {
                "category": "하부장",
                "product_name": "하부장",
                "product_code": "B-BASE-STD",
                "width_mm": 800,
                "depth_mm": 600,
                "height_mm": 850,
                "quantity": 480,
                "remarks": "주방 하부 표준 수납장"
            },
            {
                "category": "상부장",
                "product_name": "냉장고장 상부 플랩장",
                "product_code": "W-FLAP-REF",
                "width_mm": 1000,
                "depth_mm": 600,
                "height_mm": 600,
                "quantity": 120,
                "remarks": "냉장고 상부 플랩도어형 수납장"
            },
            {
                "category": "키큰장",
                "product_name": "키큰장",
                "product_code": "T-TALL-STD",
                "width_mm": 600,
                "depth_mm": 600,
                "height_mm": 2200,
                "quantity": 80,
                "remarks": "가전 기기 수납용 키큰 수납장"
            },
            {
                "category": "피라/앤드판넬",
                "product_name": "우측 마감 판넬 (비규격)",
                "product_code": "P-END-RIGHT-SPC",
                "width_mm": 211,
                "depth_mm": 600,
                "height_mm": 2200,
                "quantity": 150,
                "remarks": "우측 비규격 판넬 마감재 (211mm 가공 할증)"
            },
            {
                "category": "코니스/걸레받이",
                "product_name": "상부 마감 휠라 (코니스)",
                "product_code": "C-FILLA-TOP",
                "width_mm": 2521,
                "depth_mm": 18,
                "height_mm": 80,
                "quantity": 200,
                "remarks": "상단 코니스 휠라 (가로 2521mm 재단 할증)"
            },
            {
                "category": "기타",
                "product_name": "특수 주문형 아일랜드 식탁",
                "product_code": "CUSTOM-ISLAND",
                "width_mm": 1250,
                "depth_mm": 900,
                "height_mm": 850,
                "quantity": 60,
                "remarks": "마스터 단가 미등록. 임의 사양 조율 및 단가 확인 필요"
            }
        ]

    if output_path_override:
        output_path = output_path_override
    else:
        desktop_dir = "C:/Users/c/Desktop"
        if is_mock:
            output_filename = "디엘이앤씨_탕정_마크센텀_주방가구_예비견적서_20260602.xlsx"
        else:
            output_filename = f"디엘이앤씨_탕정_마크센텀_주방가구_예비견적서_Task_{task_id}.xlsx"
        output_path = os.path.join(desktop_dir, output_filename)

    db = SessionLocal()
    price_masters = db.query(models.CabinetPriceMaster).all()

    prices_by_code = {pm.product_code: pm for pm in price_masters if pm.product_code}
    prices_by_name = {pm.product_name: pm for pm in price_masters}
    prices_by_category = {}
    for pm in price_masters:
        if pm.category not in prices_by_category:
            prices_by_category[pm.category] = pm
        else:
            if pm.product_name == pm.category:
                prices_by_category[pm.category] = pm

    surcharge_rate = float(os.getenv("SURCHARGE_RATE", "0.30"))

    # 4. Resolve prices for each item
    estimate_items = []
    subtotal = 0

    for idx, it in enumerate(raw_items, 1):
        p_code = it["product_code"]
        p_name = it["product_name"]
        p_cat = it["category"]

        unit_price = 0
        price_source = "not_found"
        price_confidence = 0.0
        pricing_remarks = ""

        # Resolve from DB
        if p_code in prices_by_code:
            pm = prices_by_code[p_code]
            unit_price = pm.unit_price
            price_source = "exact_code"
            price_confidence = 1.0
            pricing_remarks = f"마스터 단가 코드 매치: {pm.product_code}"
        elif p_name in prices_by_name:
            pm = prices_by_name[p_name]
            unit_price = pm.unit_price
            price_source = "exact_name"
            price_confidence = 1.0
            pricing_remarks = f"마스터 품명 매치: {pm.product_name}"
        elif p_cat in prices_by_category:
            pm = prices_by_category[p_cat]
            unit_price = pm.unit_price
            price_source = "category_fallback"
            price_confidence = 0.70
            pricing_remarks = f"카테고리 대체 단가 매치 ({pm.product_name}): {pm.unit_price:,}원"
        else:
            unit_price = 0
            price_source = "not_found"
            price_confidence = 0.0
            pricing_remarks = "매칭 단가 없음 (단가 확인 필요)"

        # Non-standard width surcharge check
        is_special = False
        width = it["width_mm"]
        if width > 0 and width % 100 != 0:
            is_special = True

        if is_special:
            if unit_price > 0:
                unit_price = int(unit_price * (1.0 + surcharge_rate))
                pricing_remarks += f" (비규격 할증 {(surcharge_rate * 100):.0f}% 반영)"
            else:
                pricing_remarks += " (비규격 품목)"

        sum_price = it["quantity"] * unit_price
        subtotal += sum_price

        spec_str = f"{width} * {it['depth_mm']} * {it['height_mm']}"
        needs_review = is_special or (price_confidence < 0.80) or (unit_price == 0)

        estimate_items.append({
            "item_no": idx,
            "category": p_cat,
            "item_name": p_name,
            "spec": spec_str,
            "qty": it["quantity"],
            "unit": "EA",
            "unit_price": unit_price,
            "sum_price": sum_price,
            "price_source": price_source,
            "confidence": price_confidence,
            "needs_manual_review": needs_review,
            "remarks": it["remarks"],
            "pricing_remarks": pricing_remarks
        })

    db.close()

    # Fees and VAT
    contingency = 5000000       # 예비비 ₩5,000,000
    installation_fee = 8000000  # 설치비 ₩8,000,000
    transport_fee = 4500000     # 운반비 ₩4,500,000

    total_amount = subtotal + contingency + installation_fee + transport_fee
    vat_amount = int(total_amount * 0.10)
    grand_total = total_amount + vat_amount

    # 5. Build Excel Workbook
    wb = openpyxl.Workbook()

    # Colors
    navy_header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    soft_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    gray_zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    warning_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    accent_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    # Fonts
    font_title = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
    font_subtitle = Font(name="맑은 고딕", size=10, italic=True, color="595959")
    font_bold = Font(name="맑은 고딕", size=10, bold=True)
    font_regular = Font(name="맑은 고딕", size=10)
    font_white_bold = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    font_red_bold = Font(name="맑은 고딕", size=10, bold=True, color="FF0000")

    # Borders
    thin_line = Side(border_style="thin", color="D9D9D9")
    double_line = Side(border_style="double", color="1F497D")
    thick_line = Side(border_style="medium", color="1F497D")

    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
    border_header = Border(left=thin_line, right=thin_line, top=thick_line, bottom=thick_line)
    border_total = Border(top=thin_line, bottom=double_line)

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # --- Sheet 1: 요약 (Summary) ---
    ws1 = wb.active
    ws1.title = "요약"
    ws1.views.sheetView[0].showGridLines = True

    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 25
    ws1.column_dimensions["E"].width = 15

    # Title
    ws1["B2"] = "[샘플 견적 예시] 디엘이앤씨 탕정 마크센텀 주방가구 예비 견적서" if is_mock else f"{project_name} 예비 견적서"
    ws1["B2"].font = font_title
    ws1["B3"] = "DRAFT ESTIMATION FOR REVIEW (발주서 미작성 현장)" if is_mock else "ESTIMATION REPORT FOR REVIEW"
    ws1["B3"].font = font_subtitle

    # General Info
    ws1["B5"] = "견적 현장명:"
    ws1["B5"].font = font_bold
    ws1["C5"] = "[샘플] 디엘이앤씨 탕정 마크센텀" if is_mock else project_name
    ws1["C5"].font = font_regular

    ws1["B6"] = "원본 도면:"
    ws1["B6"].font = font_bold
    ws1["C6"] = os.path.basename(dwg_path)
    ws1["C6"].font = font_regular

    ws1["B7"] = "도면 검증 상태:"
    ws1["B7"].font = font_bold
    ws1["C7"] = f"{dwg_info} | {file_size_mb:.2f} MB ({limit_status})"
    ws1["C7"].font = font_regular

    ws1["B8"] = "견적 작성일:"
    ws1["B8"].font = font_bold
    ws1["C8"] = datetime.date.today().strftime("%Y년 %m월 %d일")
    ws1["C8"].font = font_regular

    ws1["B9"] = "견적 상태:"
    ws1["B9"].font = font_bold
    ws1["C9"] = "DRAFT / NEEDS_REVIEW (검토 요망)"
    ws1["C9"].font = font_red_bold

    # Financial Summary Table Header
    ws1.merge_cells("B11:E11")
    ws1["B11"] = "견적 금액 요약표"
    ws1["B11"].font = font_white_bold
    ws1["B11"].fill = navy_header_fill
    ws1["B11"].alignment = align_center

    summary_rows = [
        ("품목 공급가액 소계", subtotal, "각 실별 가구 품목 표준 단가 합계액 (할증 포함)"),
        ("비규격/할증 반영액", sum(i["sum_price"] for i in estimate_items if i["needs_manual_review"] and i["unit_price"] > 0) - sum(it["quantity"] * (prices_by_code.get(it["product_code"]).unit_price if it["product_code"] in prices_by_code else 0) for it in raw_items if it["width_mm"] % 100 != 0 and it["product_code"] in prices_by_code), "비규격 치수에 따른 할증 가산 누계"),
        ("설치비 (시공비)", installation_fee, "현장 설치 시공 총 비용 (추정)"),
        ("운반비 (물류비)", transport_fee, "공장 배송 및 하차 비용 (추정)"),
        ("예비비 (Contingency)", contingency, "도면 불확실성 대응용 가산 예비비"),
        ("공급가액 총계", total_amount, "소계 + 설치비 + 운반비 + 예비비"),
        ("부가가치세 (VAT 10%)", vat_amount, "공급가액 총계의 10%"),
        ("최종 견적합계액", grand_total, "공급가액 총계 + VAT (최종 청구합계)")
    ]

    for r_idx, (label, val, desc) in enumerate(summary_rows, 12):
        ws1.cell(row=r_idx, column=2, value=label).font = font_bold if "총계" in label or "합계액" in label else font_regular
        ws1.cell(row=r_idx, column=2).border = border_cell

        val_cell = ws1.cell(row=r_idx, column=3, value=val)
        val_cell.font = font_bold if "총계" in label or "합계액" in label else font_regular
        val_cell.number_format = "₩#,##0"
        val_cell.alignment = align_right
        val_cell.border = border_cell
        if "합계액" in label:
            val_cell.fill = accent_green_fill
            val_cell.font = Font(name="맑은 고딕", size=10, bold=True, color="1F497D")

        ws1.cell(row=r_idx, column=4, value=desc).font = font_subtitle
        ws1.cell(row=r_idx, column=4).border = border_cell
        ws1.merge_cells(start_row=r_idx, start_column=4, end_row=r_idx, end_column=5)

    # Warnings Block
    ws1.merge_cells("B21:E24")
    warning_text = (
        "⚠️ [주의 및 한계사항 알림]\n"
        "1. 본 견적서는 현장의 공식 '발주서(XLSX)'가 입수되지 않은 시점의 도면 분석 예비 검토 자료입니다.\n"
        "2. 현재 솔루션은 실제 CAD Geometry 및 선 분석을 통한 정밀 추출 단계가 아닌 데모/스텁(Stub) 연동 중입니다.\n"
        "3. 아일랜드 식탁 등 마스터 단가에 미등록된 품목은 단가가 0원으로 처리되어 있으며, 수동 검증이 필수적입니다.\n"
        "4. 비규격 가로 치수(211mm, 2521mm)에 대해서는 기본 단가의 30% 할증률이 자동 적용되어 가격 보정되었습니다."
    )
    ws1["B21"] = warning_text
    ws1["B21"].alignment = Alignment(wrap_text=True, vertical="top")
    ws1["B21"].fill = warning_yellow_fill
    ws1["B21"].font = Font(name="맑은 고딕", size=9, color="7F6000")

    # Apply thin border to merged warning box
    for r in range(21, 25):
        for c in range(2, 6):
            ws1.cell(row=r, column=c).border = Border(
                top=thin_line if r==21 else None,
                bottom=thin_line if r==24 else None,
                left=thin_line if c==2 else None,
                right=thin_line if c==5 else None
            )

    # --- Sheet 2: 품목 상세 (Detailed Items) ---
    ws2 = wb.create_sheet(title="품목 상세")
    ws2.views.sheetView[0].showGridLines = True

    headers = [
        "순번", "구분 (Category)", "제품명 (Item Name)", "규격 (Spec W*D*H)",
        "수량", "단위", "단가 (Unit Price)", "금액 (Sum Price)",
        "단가출처", "신뢰도", "검토요망", "특기사항"
    ]

    # Write Headers
    for c_idx, h in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=c_idx, value=h)
        cell.font = font_white_bold
        cell.fill = navy_header_fill
        cell.alignment = align_center
        cell.border = border_header

    # Write Data
    for r_offset, it in enumerate(estimate_items, 3):
        ws2.cell(row=r_offset, column=1, value=it["item_no"]).alignment = align_center
        ws2.cell(row=r_offset, column=2, value=it["category"]).alignment = align_center
        ws2.cell(row=r_offset, column=3, value=it["item_name"]).alignment = align_left
        ws2.cell(row=r_offset, column=4, value=it["spec"]).alignment = align_center

        qty_cell = ws2.cell(row=r_offset, column=5, value=it["qty"])
        qty_cell.alignment = align_right
        qty_cell.number_format = "#,##0"

        ws2.cell(row=r_offset, column=6, value=it["unit"]).alignment = align_center

        up_cell = ws2.cell(row=r_offset, column=7, value=it["unit_price"])
        up_cell.alignment = align_right
        up_cell.number_format = "₩#,##0"

        sp_cell = ws2.cell(row=r_offset, column=8, value=it["sum_price"])
        sp_cell.alignment = align_right
        sp_cell.number_format = "₩#,##0"

        ws2.cell(row=r_offset, column=9, value=it["price_source"]).alignment = align_center

        conf_cell = ws2.cell(row=r_offset, column=10, value=it["confidence"])
        conf_cell.alignment = align_right
        conf_cell.number_format = "0.0%"

        rev_cell = ws2.cell(row=r_offset, column=11, value="검토필요" if it["needs_manual_review"] else "정상")
        rev_cell.alignment = align_center
        if it["needs_manual_review"]:
            rev_cell.font = font_red_bold
            rev_cell.fill = warning_yellow_fill

        rem_cell = ws2.cell(row=r_offset, column=12, value=f"{it['remarks']} | {it['pricing_remarks']}")
        rem_cell.alignment = align_left

        # Zebra striping & borders
        for c_idx in range(1, 13):
            c_cell = ws2.cell(row=r_offset, column=c_idx)
            c_cell.border = border_cell
            c_cell.font = font_regular if c_idx != 11 or not it["needs_manual_review"] else font_red_bold
            if r_offset % 2 == 0 and not (c_idx == 11 and it["needs_manual_review"]):
                c_cell.fill = gray_zebra_fill

    # Total row at bottom
    tot_row = len(estimate_items) + 3
    ws2.cell(row=tot_row, column=3, value="품목 공급가액 합계").font = font_bold
    ws2.cell(row=tot_row, column=3).alignment = align_left

    tot_sp = ws2.cell(row=tot_row, column=8, value=f"=SUM(H3:H{tot_row-1})")
    tot_sp.font = font_bold
    tot_sp.alignment = align_right
    tot_sp.number_format = "₩#,##0"

    for c_idx in range(1, 13):
        ws2.cell(row=tot_row, column=c_idx).border = Border(top=thin_line, bottom=double_line)

    # Autoclose column widths
    for col in ws2.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format and ('₩' in cell.number_format or '%' in cell.number_format):
                max_len = max(max_len, 14)
            else:
                max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # --- Sheet 3: 검토 필요 항목 (Review Needed) ---
    ws3 = wb.create_sheet(title="검토 필요 항목")
    ws3.views.sheetView[0].showGridLines = True

    # Headers
    review_headers = ["순번", "구분", "제품명", "규격", "수량", "단가", "검토 요망 사유"]
    for c_idx, h in enumerate(review_headers, 1):
        cell = ws3.cell(row=2, column=c_idx, value=h)
        cell.font = font_white_bold
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Dark Red for Warnings
        cell.alignment = align_center
        cell.border = border_header

    # Write review items
    rev_idx = 3
    for it in estimate_items:
        if it["needs_manual_review"]:
            ws3.cell(row=rev_idx, column=1, value=it["item_no"]).alignment = align_center
            ws3.cell(row=rev_idx, column=2, value=it["category"]).alignment = align_center
            ws3.cell(row=rev_idx, column=3, value=it["item_name"]).alignment = align_left
            ws3.cell(row=rev_idx, column=4, value=it["spec"]).alignment = align_center

            qty_c = ws3.cell(row=rev_idx, column=5, value=it["qty"])
            qty_c.alignment = align_right
            qty_c.number_format = "#,##0"

            up_c = ws3.cell(row=rev_idx, column=6, value=it["unit_price"])
            up_c.alignment = align_right
            up_c.number_format = "₩#,##0"

            # Determine reason
            reason = ""
            if it["unit_price"] == 0:
                reason = "❌ 마스터 단가표 단가 누락 (등록 필요)"
            elif "할증" in it["pricing_remarks"]:
                reason = "⚠️ 비규격 가로 치수 가공 할증 적용 (실측 확인 필요)"
            else:
                reason = f"⚠️ 신뢰도 낮음 ({(it['confidence']*100):.0f}%)"

            ws3.cell(row=rev_idx, column=7, value=reason).alignment = align_left
            ws3.cell(row=rev_idx, column=7).font = font_bold

            for c_idx in range(1, 8):
                ws3.cell(row=rev_idx, column=c_idx).border = border_cell
                ws3.cell(row=rev_idx, column=c_idx).fill = warning_yellow_fill

            rev_idx += 1

    # Autofit column widths
    for col in ws3.columns:
        max_len = 0
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')))
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # --- Sheet 4: 산정 기준 및 가정 (Pricing Rules) ---
    ws4 = wb.create_sheet(title="산정 기준 및 가정")
    ws4.views.sheetView[0].showGridLines = True

    ws4.column_dimensions["B"].width = 25
    ws4.column_dimensions["C"].width = 15
    ws4.column_dimensions["D"].width = 40

    ws4["B2"] = "단가 및 견적 산정 기준"
    ws4["B2"].font = font_title

    rules = [
        ("부가가치세 (VAT)", "10%", "모든 품목 및 용역 총액에 대해 일괄 10% 과세 적용"),
        ("비규격 할증률", "30%", "가로 치수가 100mm의 배수가 아닌 경우 가공 난이도로 인한 30% 할증 적용"),
        ("기본 예비비 (Contingency)", "₩5,000,000", "도면 미확정 및 발주서 미작성에 따른 위험 요소 대응 예비비"),
        ("표준 설치 시공비", "₩8,000,000", "상/하부장 및 판넬류 전수 시공비용 추정치"),
        ("표준 운반 물류비", "₩4,500,000", "제조 공장(인천/화성 기준)에서 탕정 현장까지의 차량 배송비"),
        ("단가 매칭 우선순위", "1순위: product_code\n2순위: product_name\n3순위: category_fallback", "데이터베이스 CabinetPriceMaster 테이블 기준 매핑 규칙")
    ]

    ws4.cell(row=4, column=2, value="항목").font = font_white_bold
    ws4.cell(row=4, column=2).fill = navy_header_fill
    ws4.cell(row=4, column=2).border = border_header
    ws4.cell(row=4, column=2).alignment = align_center

    ws4.cell(row=4, column=3, value="적용값").font = font_white_bold
    ws4.cell(row=4, column=3).fill = navy_header_fill
    ws4.cell(row=4, column=3).border = border_header
    ws4.cell(row=4, column=3).alignment = align_center

    ws4.cell(row=4, column=4, value="설명").font = font_white_bold
    ws4.cell(row=4, column=4).fill = navy_header_fill
    ws4.cell(row=4, column=4).border = border_header
    ws4.cell(row=4, column=4).alignment = align_center

    for idx, (h, val, desc) in enumerate(rules, 5):
        ws4.cell(row=idx, column=2, value=h).font = font_bold
        ws4.cell(row=idx, column=2).border = border_cell
        ws4.cell(row=idx, column=2).alignment = align_left

        val_c = ws4.cell(row=idx, column=3, value=val)
        val_c.font = font_regular
        val_c.border = border_cell
        val_c.alignment = align_center

        ws4.cell(row=idx, column=4, value=desc).font = font_regular
        ws4.cell(row=idx, column=4).border = border_cell
        ws4.cell(row=idx, column=4).alignment = align_left
        ws4.cell(row=idx, column=4).alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    print(f"Quotation successfully saved to: {output_path}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Generate draft quotation Excel sheet.")
    parser.add_argument("--task-id", type=int, default=None, help="CAD task ID to query from DB")
    parser.add_argument("--output", type=str, default=None, help="Override output Excel path")
    args = parser.parse_args()

    generate_quotation(task_id=args.task_id, output_path_override=args.output)
