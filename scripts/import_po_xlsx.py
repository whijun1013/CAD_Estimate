import openpyxl
import os
import re
import argparse
import sys
import logging
from datetime import datetime

# Setup paths to import models and database
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database import SessionLocal
from models import Project, ApartmentType, MaterialSpecification, HardwareSpecification, CabinetBOM, BuildingQuantity

# Configure logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# Excel Constants
SHEET_PROJECT_INFO = "현장정보"
SHEET_SPEC_PREFIX = "사양서"
SHEET_BOM_PREFIX = "내역서"
SHEET_DONG_PREFIX = "동정보"

INFO_SCAN_MAX_ROW = 40
INFO_SCAN_MAX_COL = 10
APT_TYPE_START_ROW = 10
APT_TYPE_END_ROW = 30
APT_TYPE_COL_CODE = 11
APT_TYPE_COL_NAME = 12
APT_TYPE_COL_QTY = 13
APT_TYPE_COL_CHANGE = 14

SPEC_START_ROW = 4
SPEC_MAX_ROW = 100
SPEC_COL_PART = 4
SPEC_COL_THICKNESS = 6
SPEC_COL_GRADE = 7
SPEC_COL_MATERIAL = 8
SPEC_COL_FINISH = 9
SPEC_COL_GRAIN = 10
SPEC_COL_PRIMARY_MAT = 11
SPEC_COL_PRIMARY_MAT_DTL = 12
SPEC_COL_BACKING = 13
SPEC_COL_BACKING_DTL = 14
SPEC_COL_EDGE = 15
SPEC_COL_EDGE_DTL = 16

HW_START_ROW = 8
HW_MAX_ROW = 100
HW_COL_GRP = 19
HW_COL_NAME = 20
HW_COL_APP = 21
HW_COL_REM = 22

BOM_START_ROW = 4
BOM_MAX_ROW = 250

DONG_START_ROW = 8
DONG_MAX_ROW = 250
DONG_START_COL = 14

def parse_date(date_str):
    if not date_str:
        return None
    date_str = str(date_str).strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None

def clean_int(val, context=""):
    if val is None:
        return 0
    if isinstance(val, int):
        return val
    s = str(val).strip().replace(",", "")
    if not s:
        return 0
    try:
        return int(s)
    except ValueError:
        try:
            return int(float(s))
        except ValueError:
            return 0

def parse_po_xlsx(excel_path):
    """
    Parses PO Excel file and returns a structured dictionary representation of the data.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")

    logging.info("Opening Excel file for parsing: %s", excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 1. Parse Project Info from '현장정보'
    sheet_info = wb[SHEET_PROJECT_INFO]
    info_dict = {}

    for r in range(1, INFO_SCAN_MAX_ROW + 1):
        for c in range(1, INFO_SCAN_MAX_COL + 1):
            cell_val = sheet_info.cell(row=r, column=c).value
            if cell_val:
                cell_val = str(cell_val).strip()
                if cell_val in ("거래선", "현장명", "품목", "현장  주소", "현장담당자", "현장연락처",
                                "시공거래선", "시공담당자", "시공연락처", "최초투입일", "입주/오픈예정일",
                                "현장종류", "최고층높이", "분절공사", "PO번호", "계약번호", "작성완료일"):
                    val = sheet_info.cell(row=r, column=c+1).value
                    if val is None or val == "":
                        val = sheet_info.cell(row=r, column=c+2).value
                        if val is None or val == "":
                            val = sheet_info.cell(row=r, column=c+3).value

                    if val is not None:
                        info_dict[cell_val] = str(val).strip()

    po_number = info_dict.get("PO번호", "26-2603-0003-01")

    project_data = {
        "po_number": po_number,
        "contract_number": info_dict.get("계약번호"),
        "name": info_dict.get("현장명", "김해삼계푸르지오"),
        "client": info_dict.get("거래선"),
        "partner_installer": info_dict.get("시공거래선"),
        "item_type": info_dict.get("품목"),
        "address": info_dict.get("현장  주소"),
        "manager_name": info_dict.get("현장담당자"),
        "manager_contact": info_dict.get("현장연락처"),
        "installer_name": info_dict.get("시공담당자"),
        "installer_contact": info_dict.get("시공연락처"),
        "first_delivery_date": parse_date(info_dict.get("최초투입일")),
        "opening_date": parse_date(info_dict.get("입주/오픈예정일")),
        "site_type": info_dict.get("현장종류"),
        "max_floor": clean_int(info_dict.get("최고층높이"), "최고층높이"),
        "is_divided_work": (info_dict.get("분절공사") == "Y"),
    }

    # Parse Apartment Types from '현장정보'
    apartment_types = []
    for r in range(APT_TYPE_START_ROW, APT_TYPE_END_ROW + 1):
        t_code = sheet_info.cell(row=r, column=APT_TYPE_COL_CODE).value
        t_name = sheet_info.cell(row=r, column=APT_TYPE_COL_NAME).value
        t_qty = sheet_info.cell(row=r, column=APT_TYPE_COL_QTY).value
        t_change = sheet_info.cell(row=r, column=APT_TYPE_COL_CHANGE).value

        if t_code and t_name and str(t_code).strip() != "합계":
            t_code = str(t_code).strip()
            t_name = str(t_name).strip()
            qty = clean_int(t_qty, f"Apt Type Qty {t_code}")
            is_changed = (str(t_change).strip() == "Y")

            apartment_types.append({
                "type_name": t_code,
                "household_count": qty,
                "is_changed": is_changed,
                "specs": [],
                "hardware_specs": [],
                "boms": []
            })

    # For each apartment type, parse Specifications, BOM and building quantities
    for apt_type in apartment_types:
        t_code = apt_type["type_name"]

        # 1. Spec Sheet
        spec_sheet_name = f"{SHEET_SPEC_PREFIX}({t_code})"
        if spec_sheet_name in wb.sheetnames:
            spec_sheet = wb[spec_sheet_name]
            current_category = ""
            for r in range(SPEC_START_ROW, SPEC_MAX_ROW + 1):
                part_val = spec_sheet.cell(row=r, column=SPEC_COL_PART).value
                if part_val:
                    part_val = str(part_val).strip()
                    if part_val.startswith("[[") and part_val.endswith("]]"):
                        current_category = part_val.replace("[[", "").replace("]]", "")
                        continue
                    if current_category and part_val in ("몸통", "문짝", "코니스", "앤드판넬", "휠라", "걸레받이"):
                        thickness = spec_sheet.cell(row=r, column=SPEC_COL_THICKNESS).value
                        grade = spec_sheet.cell(row=r, column=SPEC_COL_GRADE).value
                        material = spec_sheet.cell(row=r, column=SPEC_COL_MATERIAL).value
                        finish_method = spec_sheet.cell(row=r, column=SPEC_COL_FINISH).value
                        grain_direction = spec_sheet.cell(row=r, column=SPEC_COL_GRAIN).value
                        primary_material = spec_sheet.cell(row=r, column=SPEC_COL_PRIMARY_MAT).value
                        primary_detail = spec_sheet.cell(row=r, column=SPEC_COL_PRIMARY_MAT_DTL).value
                        backing_material = spec_sheet.cell(row=r, column=SPEC_COL_BACKING).value
                        backing_detail = spec_sheet.cell(row=r, column=SPEC_COL_BACKING_DTL).value
                        edge_material = spec_sheet.cell(row=r, column=SPEC_COL_EDGE).value
                        edge_detail = spec_sheet.cell(row=r, column=SPEC_COL_EDGE_DTL).value

                        apt_type["specs"].append({
                            "category": current_category,
                            "part_name": part_val,
                            "thickness": str(thickness).strip() if thickness else None,
                            "grade": str(grade).strip() if grade else None,
                            "material": str(material).strip() if material else None,
                            "finish_method": str(finish_method).strip() if finish_method else None,
                            "grain_direction": str(grain_direction).strip() if grain_direction else None,
                            "primary_material": str(primary_material).strip() if primary_material else None,
                            "primary_material_detail": str(primary_detail).strip() if primary_detail else None,
                            "backing_material": str(backing_material).strip() if backing_material else None,
                            "backing_material_detail": str(backing_detail).strip() if backing_detail else None,
                            "edge_material": str(edge_material).strip() if edge_material else None,
                            "edge_material_detail": str(edge_detail).strip() if edge_detail else None
                        })

            # Hardware Specs
            for r in range(HW_START_ROW, HW_MAX_ROW + 1):
                item_grp = spec_sheet.cell(row=r, column=HW_COL_GRP).value
                item_name = spec_sheet.cell(row=r, column=HW_COL_NAME).value
                app_val = spec_sheet.cell(row=r, column=HW_COL_APP).value
                rem_val = spec_sheet.cell(row=r, column=HW_COL_REM).value

                if item_name and str(item_name).strip():
                    apt_type["hardware_specs"].append({
                        "item_group": str(item_grp).strip() if item_grp else "사양",
                        "item_name": str(item_name).strip(),
                        "application": str(app_val).strip() if app_val else None,
                        "special_remarks": str(rem_val).strip() if rem_val else None
                    })

        # 2. BOM Sheet
        bom_sheet_name = f"{SHEET_BOM_PREFIX}({t_code})"
        boms_by_item_no = {}
        if bom_sheet_name in wb.sheetnames:
            bom_sheet = wb[bom_sheet_name]
            current_category = ""
            for r in range(BOM_START_ROW, BOM_MAX_ROW + 1):
                col1 = bom_sheet.cell(row=r, column=1).value
                col5_no = bom_sheet.cell(row=r, column=5).value
                col6_name = bom_sheet.cell(row=r, column=6).value

                if col1 and str(col1).strip() in ("상부장", "하부장", "키큰장", "보조주방", "보조주방1", "보조주방2", "피라/앤드판넬", "코니스/걸레받이"):
                    current_category = str(col1).strip()
                    continue

                no_val = clean_int(col5_no, f"BOM Item No row {r}")
                if no_val > 0 and col6_name:
                    status = bom_sheet.cell(row=r, column=1).value
                    is_special = bom_sheet.cell(row=r, column=4).value
                    product_code = bom_sheet.cell(row=r, column=7).value
                    attr_code = bom_sheet.cell(row=r, column=8).value

                    w = bom_sheet.cell(row=r, column=9).value
                    h = bom_sheet.cell(row=r, column=10).value
                    d = bom_sheet.cell(row=r, column=11).value

                    direction = bom_sheet.cell(row=r, column=12).value
                    qd_l = bom_sheet.cell(row=r, column=13).value
                    qd_m = bom_sheet.cell(row=r, column=14).value
                    qd_r = bom_sheet.cell(row=r, column=15).value
                    qo_l = bom_sheet.cell(row=r, column=16).value
                    qo_m = bom_sheet.cell(row=r, column=17).value
                    qo_r = bom_sheet.cell(row=r, column=18).value
                    qty_sum = bom_sheet.cell(row=r, column=19).value
                    remarks = bom_sheet.cell(row=r, column=20).value

                    bom_item = {
                        "category": current_category if current_category else "기타",
                        "status": str(status).strip() if status else None,
                        "is_special": (str(is_special).strip() == "S"),
                        "item_no": no_val,
                        "product_name": str(col6_name).strip(),
                        "product_code": str(product_code).strip() if product_code else None,
                        "attribute_code": str(attr_code).strip() if attr_code else None,
                        "width": clean_int(w, f"BOM w row {r}"),
                        "height": clean_int(h, f"BOM h row {r}"),
                        "depth": clean_int(d, f"BOM d row {r}"),
                        "base_direction": str(direction).strip() if direction else None,
                        "qty_drawing_left": clean_int(qd_l, f"BOM qd_l row {r}"),
                        "qty_drawing_mid": clean_int(qd_m, f"BOM qd_m row {r}"),
                        "qty_drawing_right": clean_int(qd_r, f"BOM qd_r row {r}"),
                        "qty_opposite_left": clean_int(qo_l, f"BOM qo_l row {r}"),
                        "qty_opposite_mid": clean_int(qo_m, f"BOM qo_m row {r}"),
                        "qty_opposite_right": clean_int(qo_r, f"BOM qo_r row {r}"),
                        "qty_sum": clean_int(qty_sum, f"BOM qty_sum row {r}"),
                        "remarks": str(remarks).strip() if remarks else None,
                        "building_quantities": []
                    }
                    apt_type["boms"].append(bom_item)
                    boms_by_item_no[(no_val, bom_item["product_name"])] = bom_item

        # 3. Dong Info Sheet
        dong_sheet_name = f"{SHEET_DONG_PREFIX}({t_code})"
        if dong_sheet_name in wb.sheetnames:
            dong_sheet = wb[dong_sheet_name]

            # Map columns to building and line numbers
            col_mappings = {}
            current_building = ""
            max_cols = dong_sheet.max_column

            for c in range(DONG_START_COL, max_cols + 1):
                header6 = dong_sheet.cell(row=6, column=c).value
                if not header6:
                    continue

                b_val = dong_sheet.cell(row=4, column=c).value
                if b_val:
                    current_building = str(b_val).strip()

                line_val = None
                for check_c in range(c, DONG_START_COL - 1, -1):
                    lv = dong_sheet.cell(row=5, column=check_c).value
                    if lv is not None:
                        line_val = str(lv).strip()
                        break

                if str(header6).strip() == "합계" and current_building and line_val:
                    col_mappings[c] = (current_building, line_val)

            for r in range(DONG_START_ROW, DONG_MAX_ROW + 1):
                col4_no = dong_sheet.cell(row=r, column=4).value
                col6_name = dong_sheet.cell(row=r, column=6).value

                no_val = clean_int(col4_no, f"Dong Row {r}")
                if no_val > 0 and col6_name:
                    p_name = str(col6_name).strip()
                    bom_item = boms_by_item_no.get((no_val, p_name))
                    if bom_item:
                        for c, (b_no, l_no) in col_mappings.items():
                            qty_val = dong_sheet.cell(row=r, column=c).value
                            qty = clean_int(qty_val, f"Dong Qty row {r} col {c}")
                            if qty > 0:
                                bom_item["building_quantities"].append({
                                    "building_no": b_no,
                                    "line_no": l_no,
                                    "qty": qty
                                })

    return {
        "project": project_data,
        "apartment_types": apartment_types
    }

def import_to_db(db, parsed_data, destructive_reload=False, prune_missing=False):
    """
    Idempotently imports parsed PO data into the database.

    Deletion policy (non-destructive by default):
    ─────────────────────────────────────────────
    destructive_reload=False (default, SAFE):
        Updates Project + ApartmentType fields in-place.
        Does NOT delete any CADTask, Quotation, QuotationItem, or Audit records.
        Apartment types / BOM rows are only added or updated, never removed
        (unless prune_missing=True is also passed).

    destructive_reload=True (DANGER):
        Deletes the entire Project row (cascade: ApartmentTypes, BOMs, CADTasks,
        Quotations, QuotationItems, Audits). Then re-imports from scratch.
        Use only when you want a full clean slate.
        ⚠️ All CADTask, Quotation, and Audit records for this project will be lost.

    prune_missing=False (default, SAFE):
        ApartmentType rows and BOM rows that are missing from the new XLSX are
        kept in the database. No records are deleted.

    prune_missing=True:
        ApartmentType rows not present in the new XLSX are deleted (cascade: BOMs,
        BuildingQuantities). BOM rows not present in the XLSX are also deleted.
        CADTask / Quotation / Audit records are NOT affected.
        Only safe to use when you are sure the XLSX reflects the full current state.

    Combining destructive_reload=True with prune_missing=True is equivalent to
    destructive_reload=True alone (project delete already removes everything).
    """
    proj_info = parsed_data["project"]
    po_number = proj_info["po_number"]

    # 1. Handle Project Idempotency
    existing_project = db.query(Project).filter(Project.po_number == po_number).first()

    if existing_project and destructive_reload:
        logging.warning(
            "DESTRUCTIVE RELOAD: Deleting project '%s' (PO: %s) and ALL related records "
            "(CADTask, Quotation, QuotationItem, Audit). This cannot be undone.",
            existing_project.name, po_number
        )
        db.delete(existing_project)
        db.commit()
        existing_project = None

    if existing_project:
        logging.info("Project %s (PO: %s) already exists. Performing NON-DESTRUCTIVE updates...", existing_project.name, po_number)
        # Update Project fields
        project = existing_project
        project.contract_number = proj_info["contract_number"]
        project.name = proj_info["name"]
        project.client = proj_info["client"]
        project.partner_installer = proj_info["partner_installer"]
        project.item_type = proj_info["item_type"]
        project.address = proj_info["address"]
        project.manager_name = proj_info["manager_name"]
        project.manager_contact = proj_info["manager_contact"]
        project.installer_name = proj_info["installer_name"]
        project.installer_contact = proj_info["installer_contact"]
        project.first_delivery_date = proj_info["first_delivery_date"]
        project.opening_date = proj_info["opening_date"]
        project.site_type = proj_info["site_type"]
        project.max_floor = proj_info["max_floor"]
        project.is_divided_work = proj_info["is_divided_work"]
        project.remarks = "Idempotently updated from scripts/import_po_xlsx.py"
        db.commit()
        db.refresh(project)
    else:
        # Create Project
        project = Project(
            po_number=po_number,
            contract_number=proj_info["contract_number"],
            name=proj_info["name"],
            client=proj_info["client"],
            partner_installer=proj_info["partner_installer"],
            item_type=proj_info["item_type"],
            address=proj_info["address"],
            manager_name=proj_info["manager_name"],
            manager_contact=proj_info["manager_contact"],
            installer_name=proj_info["installer_name"],
            installer_contact=proj_info["installer_contact"],
            first_delivery_date=proj_info["first_delivery_date"],
            opening_date=proj_info["opening_date"],
            site_type=proj_info["site_type"],
            max_floor=proj_info["max_floor"],
            is_divided_work=proj_info["is_divided_work"],
            remarks="Idempotently imported from scripts/import_po_xlsx.py"
        )
        db.add(project)
        db.commit()
        db.refresh(project)

    # Return count aggregates
    stats = {
        "created": {
            "projects": 1 if not existing_project else 0,
            "apartment_types": 0,
            "material_specs": 0,
            "hardware_specs": 0,
            "cabinet_boms": 0,
            "building_quantities": 0
        },
        "updated": {
            "projects": 1 if existing_project else 0,
            "apartment_types": 0,
            "material_specs": 0,
            "hardware_specs": 0,
            "cabinet_boms": 0,
            "building_quantities": 0
        },
        "preserved": {
            "apartment_types": 0,
            "material_specs": 0,
            "hardware_specs": 0,
            "cabinet_boms": 0,
            "building_quantities": 0
        },
        "pruned": {
            "apartment_types": 0,
            "material_specs": 0,
            "hardware_specs": 0,
            "cabinet_boms": 0,
            "building_quantities": 0
        }
    }

    diff_report = {
        "added": [],
        "removed": [],
        "qty_changed": [],
        "spec_changed": []
    }

    total_types = 0
    total_boms = 0

    # 2. Upsert Apartment Types & Child Rows
    parsed_types = parsed_data["apartment_types"]
    existing_types_dict = {t.type_name: t for t in db.query(ApartmentType).filter(ApartmentType.project_id == project.id).all()}

    active_type_ids = []

    for type_data in parsed_types:
        t_name = type_data["type_name"]
        if t_name in existing_types_dict:
            apt_type = existing_types_dict[t_name]
            apt_type.household_count = type_data["household_count"]
            apt_type.is_changed = type_data["is_changed"]
            db.commit()
            db.refresh(apt_type)
            stats["updated"]["apartment_types"] += 1
        else:
            apt_type = ApartmentType(
                project_id=project.id,
                type_name=t_name,
                household_count=type_data["household_count"],
                is_changed=type_data["is_changed"]
            )
            db.add(apt_type)
            db.commit()
            db.refresh(apt_type)
            stats["created"]["apartment_types"] += 1

        active_type_ids.append(apt_type.id)
        total_types += 1

        # 2a. Upsert Material Specifications
        existing_specs = db.query(MaterialSpecification).filter(MaterialSpecification.type_id == apt_type.id).all()
        existing_specs_dict = {(s.category, s.part_name): s for s in existing_specs}
        active_spec_ids = []

        for spec in type_data["specs"]:
            key = (spec["category"], spec["part_name"])
            if key in existing_specs_dict:
                m_spec = existing_specs_dict[key]
                m_spec.thickness = spec["thickness"]
                m_spec.grade = spec["grade"]
                m_spec.material = spec["material"]
                m_spec.finish_method = spec["finish_method"]
                m_spec.grain_direction = spec["grain_direction"]
                m_spec.primary_material = spec["primary_material"]
                m_spec.primary_material_detail = spec["primary_material_detail"]
                m_spec.backing_material = spec["backing_material"]
                m_spec.backing_material_detail = spec["backing_material_detail"]
                m_spec.edge_material = spec["edge_material"]
                m_spec.edge_material_detail = spec["edge_material_detail"]
                db.commit()
                stats["updated"]["material_specs"] += 1
            else:
                m_spec = MaterialSpecification(
                    type_id=apt_type.id,
                    category=spec["category"],
                    part_name=spec["part_name"],
                    thickness=spec["thickness"],
                    grade=spec["grade"],
                    material=spec["material"],
                    finish_method=spec["finish_method"],
                    grain_direction=spec["grain_direction"],
                    primary_material=spec["primary_material"],
                    primary_material_detail=spec["primary_material_detail"],
                    backing_material=spec["backing_material"],
                    backing_material_detail=spec["backing_material_detail"],
                    edge_material=spec["edge_material"],
                    edge_material_detail=spec["edge_material_detail"]
                )
                db.add(m_spec)
                db.commit()
                db.refresh(m_spec)
                stats["created"]["material_specs"] += 1
            active_spec_ids.append(m_spec.id)

        # Delete old specs only if prune_missing is enabled
        for s in existing_specs:
            if s.id not in active_spec_ids:
                if prune_missing:
                    db.delete(s)
                    stats["pruned"]["material_specs"] += 1
                else:
                    stats["preserved"]["material_specs"] += 1
        db.commit()

        # 2b. Upsert Hardware Specifications
        existing_hws = db.query(HardwareSpecification).filter(HardwareSpecification.type_id == apt_type.id).all()
        existing_hws_dict = {(hw.item_group, hw.item_name): hw for hw in existing_hws}
        active_hw_ids = []

        for hw in type_data["hardware_specs"]:
            key = (hw["item_group"], hw["item_name"])
            if key in existing_hws_dict:
                hw_spec = existing_hws_dict[key]
                hw_spec.application = hw["application"]
                hw_spec.special_remarks = hw["special_remarks"]
                db.commit()
                stats["updated"]["hardware_specs"] += 1
            else:
                hw_spec = HardwareSpecification(
                    type_id=apt_type.id,
                    item_group=hw["item_group"],
                    item_name=hw["item_name"],
                    application=hw["application"],
                    special_remarks=hw["special_remarks"]
                )
                db.add(hw_spec)
                db.commit()
                db.refresh(hw_spec)
                stats["created"]["hardware_specs"] += 1
            active_hw_ids.append(hw_spec.id)

        # Delete old hardware specs only if prune_missing is enabled
        for hw in existing_hws:
            if hw.id not in active_hw_ids:
                if prune_missing:
                    db.delete(hw)
                    stats["pruned"]["hardware_specs"] += 1
                else:
                    stats["preserved"]["hardware_specs"] += 1
        db.commit()

        # 2c. Upsert CabinetBOM & BuildingQuantity
        existing_boms = db.query(CabinetBOM).filter(CabinetBOM.type_id == apt_type.id).all()
        existing_boms_dict = {(bom.item_no, bom.product_name): bom for bom in existing_boms}
        active_bom_ids = []

        for bom in type_data["boms"]:
            key = (bom["item_no"], bom["product_name"])
            if key in existing_boms_dict:
                bom_item = existing_boms_dict[key]

                # Check for diffs
                old_qty = bom_item.qty_sum
                new_qty = bom["qty_sum"]
                if old_qty != new_qty:
                    diff_report["qty_changed"].append(f"[{t_name}] {bom['product_name']}: {old_qty} -> {new_qty}")

                old_w, old_h, old_d = bom_item.width, bom_item.height, bom_item.depth
                new_w, new_h, new_d = bom["width"], bom["height"], bom["depth"]
                if (old_w, old_h, old_d) != (new_w, new_h, new_d):
                    diff_report["spec_changed"].append(f"[{t_name}] {bom['product_name']}: {old_w}*{old_h}*{old_d} -> {new_w}*{new_h}*{new_d}")

                bom_item.category = bom["category"]
                bom_item.status = bom["status"]
                bom_item.is_special = bom["is_special"]
                bom_item.product_code = bom["product_code"]
                bom_item.attribute_code = bom["attribute_code"]
                bom_item.width = bom["width"]
                bom_item.height = bom["height"]
                bom_item.depth = bom["depth"]
                bom_item.base_direction = bom["base_direction"]
                bom_item.qty_drawing_left = bom["qty_drawing_left"]
                bom_item.qty_drawing_mid = bom["qty_drawing_mid"]
                bom_item.qty_drawing_right = bom["qty_drawing_right"]
                bom_item.qty_opposite_left = bom["qty_opposite_left"]
                bom_item.qty_opposite_mid = bom["qty_opposite_mid"]
                bom_item.qty_opposite_right = bom["qty_opposite_right"]
                bom_item.qty_sum = bom["qty_sum"]
                bom_item.remarks = bom["remarks"]
                db.commit()
                stats["updated"]["cabinet_boms"] += 1
            else:
                diff_report["added"].append(f"[{t_name}] {bom['product_name']} ({bom['qty_sum']} EA)")
                bom_item = CabinetBOM(
                    type_id=apt_type.id,
                    category=bom["category"],
                    status=bom["status"],
                    is_special=bom["is_special"],
                    item_no=bom["item_no"],
                    product_name=bom["product_name"],
                    product_code=bom["product_code"],
                    attribute_code=bom["attribute_code"],
                    width=bom["width"],
                    height=bom["height"],
                    depth=bom["depth"],
                    base_direction=bom["base_direction"],
                    qty_drawing_left=bom["qty_drawing_left"],
                    qty_drawing_mid=bom["qty_drawing_mid"],
                    qty_drawing_right=bom["qty_drawing_right"],
                    qty_opposite_left=bom["qty_opposite_left"],
                    qty_opposite_mid=bom["qty_opposite_mid"],
                    qty_opposite_right=bom["qty_opposite_right"],
                    qty_sum=bom["qty_sum"],
                    remarks=bom["remarks"]
                )
                db.add(bom_item)
                db.commit()
                db.refresh(bom_item)
                stats["created"]["cabinet_boms"] += 1

            active_bom_ids.append(bom_item.id)
            total_boms += 1

            # Upsert BuildingQuantity
            existing_bqs = db.query(BuildingQuantity).filter(BuildingQuantity.bom_id == bom_item.id).all()
            existing_bqs_dict = {(bq.building_no, bq.line_no): bq for bq in existing_bqs}
            active_bq_ids = []

            for bq in bom["building_quantities"]:
                bq_key = (bq["building_no"], bq["line_no"])
                if bq_key in existing_bqs_dict:
                    bq_item = existing_bqs_dict[bq_key]
                    bq_item.qty = bq["qty"]
                    db.commit()
                    stats["updated"]["building_quantities"] += 1
                else:
                    bq_item = BuildingQuantity(
                        bom_id=bom_item.id,
                        building_no=bq["building_no"],
                        line_no=bq["line_no"],
                        qty=bq["qty"]
                    )
                    db.add(bq_item)
                    db.commit()
                    db.refresh(bq_item)
                    stats["created"]["building_quantities"] += 1
                active_bq_ids.append(bq_item.id)

            # Delete old building quantities
            for bq in existing_bqs:
                if bq.id not in active_bq_ids:
                    if prune_missing:
                        db.delete(bq)
                        stats["pruned"]["building_quantities"] += 1
                    else:
                        stats["preserved"]["building_quantities"] += 1
            db.commit()

        # Delete old BOMs
        for bom in existing_boms:
            if bom.id not in active_bom_ids:
                diff_report["removed"].append(f"[{t_name}] {bom.product_name}")
                if prune_missing:
                    db.delete(bom)
                    stats["pruned"]["cabinet_boms"] += 1
                else:
                    stats["preserved"]["cabinet_boms"] += 1
        db.commit()

    # Delete old apartment types not present in new XLSX
    for type_name, t in existing_types_dict.items():
        if t.id not in active_type_ids:
            if prune_missing:
                logging.info(
                    "prune_missing=True: Removing ApartmentType '%s' (id=%d) not in new XLSX.",
                    type_name, t.id
                )
                db.delete(t)
                stats["pruned"]["apartment_types"] += 1
            else:
                stats["preserved"]["apartment_types"] += 1
    db.commit()

    return project, total_types, total_boms, stats, diff_report

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import Purchase Order Excel sheet into the CAD_Estimate database.")
    parser.add_argument("--file", required=True, help="Path to the Excel file to import. Keep real project workbooks outside public Git.")
    parser.add_argument(
        "--destructive-reload",
        action="store_true",
        help=(
            "DANGER: Delete the entire project (including CADTask, Quotation, Audit records) "
            "and reimport from scratch. Cannot be undone."
        )
    )
    parser.add_argument(
        "--prune-missing",
        action="store_true",
        help=(
            "Remove ApartmentType/BOM rows that are in the database but missing from the new XLSX. "
            "Does NOT affect CADTask, Quotation, or Audit records. "
            "Use when the XLSX is the definitive source of truth for apartment types."
        )
    )
    args = parser.parse_args()

    db = SessionLocal()
    try:
        parsed = parse_po_xlsx(args.file)
        project, num_types, num_boms, stats, diff_report = import_to_db(
            db, parsed,
            destructive_reload=args.destructive_reload,
            prune_missing=args.prune_missing
        )

        print(f"SUCCESS: Import completed successfully!")
        print(f"Project Name: {project.name}")
        print(f"P/O Number  : {project.po_number}")
        print(f"Apartment Types Imported: {num_types}")
        print(f"Cabinet BOM Items Imported: {num_boms}")
        print(f"DB Import stats: {stats}")

        print("\n--- BOM Diff Report ---")
        if diff_report["added"]:
            print(f"Added ({len(diff_report['added'])} items):")
            for item in diff_report["added"][:10]:
                print(f"  + {item}")
            if len(diff_report["added"]) > 10: print("  ... and more")

        if diff_report["removed"]:
            print(f"\nRemoved ({len(diff_report['removed'])} items):")
            for item in diff_report["removed"][:10]:
                print(f"  - {item}")
            if len(diff_report["removed"]) > 10: print("  ... and more")

        if diff_report["qty_changed"]:
            print(f"\nQuantity Changed ({len(diff_report['qty_changed'])} items):")
            for item in diff_report["qty_changed"][:10]:
                print(f"  * {item}")
            if len(diff_report["qty_changed"]) > 10: print("  ... and more")

        if diff_report["spec_changed"]:
            print(f"\nSpecification Changed ({len(diff_report['spec_changed'])} items):")
            for item in diff_report["spec_changed"][:10]:
                print(f"  * {item}")
            if len(diff_report["spec_changed"]) > 10: print("  ... and more")

        if not any(diff_report.values()):
            print("No BOM changes detected.")

    except Exception as e:
        logging.exception("Failed to run import_po_xlsx script.")
        sys.exit(1)
    finally:
        db.close()
